from flask import Blueprint, render_template, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime
from .auth import login_required, permission_required
from .database import get_db_connection

bsecp_bp = Blueprint('bsecp', __name__)

@bsecp_bp.route('/bsecp/modules')
@login_required
@permission_required('bsecp_modules')
def bsecp_modules():
    return render_template('bsecp_modules.html')

@bsecp_bp.route('/auth_query', methods=['GET', 'POST'])
@login_required
@permission_required('auth_query')
def auth_query():
    results = None
    if request.method == 'POST':
        user_id = request.form.get('user_id')
        resource = request.form.get('resource')
        
        try:
            connection = get_db_connection()
            with connection.cursor() as cursor:
                query = "SELECT * FROM auth_records WHERE 1=1"
                params = []
                
                if user_id:
                    query += " AND user_id LIKE %s"
                    params.append(f"%{user_id}%")
                
                if resource:
                    query += " AND resource LIKE %s"
                    params.append(f"%{resource}%")
                
                if params:
                    cursor.execute(query, params)
                else:
                    cursor.execute("SELECT * FROM auth_records LIMIT 50")
                
                results = cursor.fetchall()
                
                for result in results:
                    if result.get('grant_time'):
                        result['grant_time'] = result['grant_time'].strftime('%Y-%m-%d %H:%M:%S')
                        
            connection.close()
        except Exception as e:
            print(f"查询授权记录时出错: {e}")
    
    return render_template('auth_query.html', results=results)

@bsecp_bp.route('/bsecp/authorization-query')
@login_required
@permission_required('auth_query')
def bsecp_authorization_query():
    return render_template('bsecp_authorization_query.html')

@bsecp_bp.route('/bsecp/auth-query-modify')
@login_required
@permission_required('auth_detail')
def bsecp_auth_query_modify():
    return render_template('bsecp_auth_query_modify.html')

@bsecp_bp.route('/api/bsecp/auth-query/customer')
@login_required
@permission_required('auth_query')
def get_customer_info():
    try:
        customer_name = request.args.get('name', '')
        
        if not customer_name:
            return jsonify({'success': False, 'message': '客户名称不能为空', 'customers': []})
        
        connection = get_db_connection()
        with connection.cursor() as cursor:
            query = "SELECT * FROM cljc_customer WHERE CU_NAME LIKE %s"
            cursor.execute(query, (f'%{customer_name}%',))
            customers = cursor.fetchall()
        connection.close()
        
        return jsonify({
            'success': True,
            'customers': customers
        })
        
    except Exception as e:
        print(f"获取客户信息时出错: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e),
            'customers': []
        })

@bsecp_bp.route('/api/bsecp/auth-query/customer-suggestions')
@login_required
@permission_required('auth_query')
def get_customer_suggestions():
    try:
        keyword = request.args.get('keyword', '')
        limit = int(request.args.get('limit', 10))
        
        if not keyword:
            return jsonify({'success': False, 'message': '关键字不能为空', 'suggestions': []})
        
        connection = get_db_connection()
        with connection.cursor() as cursor:
            query = """
                SELECT CU_ID, CU_NAME, CU_CODE, CU_ORGCODE 
                FROM cljc_customer 
                WHERE CU_NAME LIKE %s 
                ORDER BY CU_NAME 
                LIMIT %s
            """
            cursor.execute(query, (f'%{keyword}%', limit))
            suggestions = cursor.fetchall()
        connection.close()
        
        print(f"搜索建议查询成功: keyword={keyword}, limit={limit}, count={len(suggestions)}")
        
        return jsonify({
            'success': True,
            'suggestions': suggestions
        })
        
    except Exception as e:
        print(f"获取客户建议时出错: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e),
            'suggestions': []
        })

@bsecp_bp.route('/api/bsecp/auth-query/license')
@login_required
@permission_required('auth_query')
def get_license_info():
    try:
        customer_name = request.args.get('name', '')
        
        if not customer_name:
            return jsonify({'success': False, 'message': '客户名称不能为空', 'licenses': []})
        
        connection = get_db_connection()
        with connection.cursor() as cursor:
            query = "SELECT * FROM cljc_license WHERE LIC_CUNAME=%s"
            cursor.execute(query, (customer_name,))
            licenses = cursor.fetchall()
        connection.close()
        
        return jsonify({
            'success': True,
            'licenses': licenses
        })
        
    except Exception as e:
        print(f"获取授权信息时出错: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e),
            'licenses': []
        })

@bsecp_bp.route('/api/bsecp/auth-query/license-detail')
@login_required
@permission_required('auth_query')
def get_license_detail_info():
    try:
        customer_name = request.args.get('name', '')
        
        if not customer_name:
            return jsonify({'success': False, 'message': '客户名称不能为空', 'licenseDetails': []})
        
        connection = get_db_connection()
        with connection.cursor() as cursor:
            query = "SELECT * FROM cljc_licensedetail WHERE LICD_SRCCUNAME=%s"
            cursor.execute(query, (customer_name,))
            license_details = cursor.fetchall()
        connection.close()
        
        return jsonify({
            'success': True,
            'licenseDetails': license_details
        })
        
    except Exception as e:
        print(f"获取授权详情信息时出错: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e),
            'licenseDetails': []
        })

@bsecp_bp.route('/api/bsecp/modules')
@login_required
@permission_required('bsecp_modules')
def get_bsecp_modules():
    try:
        search_name = request.args.get('name', '')
        search_code = request.args.get('code', '')
        search_status = request.args.get('status', '')
        
        connection = get_db_connection()
        with connection.cursor() as cursor:
            query = "SELECT * FROM cljc_module WHERE 1=1"
            params = []
            
            if search_name:
                query += " AND MD_NAME LIKE %s"
                params.append(f"%{search_name}%")
                
            if search_code:
                query += " AND MD_CODE LIKE %s"
                params.append(f"%{search_code}%")
                
            if search_status:
                query += " AND MD_STATE = %s"
                params.append(int(search_status))
            
            query += " ORDER BY MD_CREATE_DATE DESC"
            cursor.execute(query, params)
            modules = cursor.fetchall()
        connection.close()
        
        for module in modules:
            if module['MD_CREATE_DATE'] and hasattr(module['MD_CREATE_DATE'], 'strftime'):
                module['MD_CREATE_DATE'] = module['MD_CREATE_DATE'].strftime('%Y-%m-%d %H:%M:%S')
            if module['MD_MODIFY_DATE'] and hasattr(module['MD_MODIFY_DATE'], 'strftime'):
                module['MD_MODIFY_DATE'] = module['MD_MODIFY_DATE'].strftime('%Y-%m-%d %H:%M:%S')
            if module['MD_FORBIT_DATE'] and hasattr(module['MD_FORBIT_DATE'], 'strftime'):
                module['MD_FORBIT_DATE'] = module['MD_FORBIT_DATE'].strftime('%Y-%m-%d %H:%M:%S')
        
        return jsonify({
            'success': True,
            'modules': modules,
            'total': len(modules)
        })
        
    except Exception as e:
        print(f"获取BSECP模块数据时出错: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e),
            'modules': []
        })

@bsecp_bp.route('/api/bsecp/authorization-records')
@login_required
@permission_required('auth_query')
def get_bsecp_authorization_records():
    try:
        search_term = request.args.get('search', '')
        
        connection = get_db_connection()
        with connection.cursor() as cursor:
            query = "SELECT * FROM OrderAutoAuthorizationQueue WHERE 1=1"
            params = []
            
            if search_term:
                query += " AND (OD_SERIAL_NUMBER LIKE %s OR OD_CONTRACT_NUMBER LIKE %s OR OD_BMPID LIKE %s)"
                params.extend([f"%{search_term}%", f"%{search_term}%", f"%{search_term}%"])
            
            query += " ORDER BY CreateTime DESC"
            cursor.execute(query, params)
            records = cursor.fetchall()
        connection.close()
        
        for record in records:
            if record['CreateTime'] and hasattr(record['CreateTime'], 'strftime'):
                record['CreateTime'] = record['CreateTime'].strftime('%Y-%m-%d %H:%M:%S')
            if record['AutoAuthHandleTime'] and hasattr(record['AutoAuthHandleTime'], 'strftime'):
                record['AutoAuthHandleTime'] = record['AutoAuthHandleTime'].strftime('%Y-%m-%d %H:%M:%S')
        
        return jsonify({
            'success': True,
            'records': records,
            'total': len(records)
        })
        
    except Exception as e:
        print(f"获取BSECP授权记录数据时出错: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e),
            'records': []
        })

@bsecp_bp.route('/api/bsecp/authorization-records/export')
@login_required
@permission_required('auth_query')
def export_bsecp_authorization_records():
    try:
        connection = get_db_connection()
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM OrderAutoAuthorizationQueue ORDER BY CreateTime DESC")
            records = cursor.fetchall()
        connection.close()
        
        for record in records:
            if record['CreateTime'] and hasattr(record['CreateTime'], 'strftime'):
                record['CreateTime'] = record['CreateTime'].strftime('%Y-%m-%d %H:%M:%S')
            if record['AutoAuthHandleTime'] and hasattr(record['AutoAuthHandleTime'], 'strftime'):
                record['AutoAuthHandleTime'] = record['AutoAuthHandleTime'].strftime('%Y-%m-%d %H:%M:%S')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "BSECP授权记录数据"
        
        headers = [
            "ID", "订单号", "合同号", "BMPID", "授权状态", "备注", 
            "创建时间", "处理时间", "处理结果", "结果描述"
        ]
        
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        for row, record in enumerate(records, 2):
            ws.cell(row=row, column=1, value=record['FId'])
            ws.cell(row=row, column=2, value=record['OD_SERIAL_NUMBER'])
            ws.cell(row=row, column=3, value=record['OD_CONTRACT_NUMBER'])
            ws.cell(row=row, column=4, value=record['OD_BMPID'])
            ws.cell(row=row, column=5, value=record['AutoAuthFlag'])
            ws.cell(row=row, column=6, value=record['Remark'])
            ws.cell(row=row, column=7, value=record['CreateTime'])
            ws.cell(row=row, column=8, value=record['AutoAuthHandleTime'])
            ws.cell(row=row, column=9, value=record['AutoAuthHandleResult'])
            ws.cell(row=row, column=10, value=record['AutoAuthHandleResultDesc'])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"BSECP授权记录导出-{timestamp}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        print(f"导出BSECP授权记录数据时出错: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': '导出失败'})

@bsecp_bp.route('/api/bsecp/modules/export')
@login_required
@permission_required('bsecp_modules')
def export_bsecp_modules():
    try:
        connection = get_db_connection()
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM cljc_module ORDER BY MD_CREATE_DATE DESC")
            modules = cursor.fetchall()
        connection.close()
        
        for module in modules:
            if module['MD_CREATE_DATE'] and hasattr(module['MD_CREATE_DATE'], 'strftime'):
                module['MD_CREATE_DATE'] = module['MD_CREATE_DATE'].strftime('%Y-%m-%d %H:%M:%S')
            if module['MD_MODIFY_DATE'] and hasattr(module['MD_MODIFY_DATE'], 'strftime'):
                module['MD_MODIFY_DATE'] = module['MD_MODIFY_DATE'].strftime('%Y-%m-%d %H:%M:%S')
            if module['MD_FORBIT_DATE'] and hasattr(module['MD_FORBIT_DATE'], 'strftime'):
                module['MD_FORBIT_DATE'] = module['MD_FORBIT_DATE'].strftime('%Y-%m-%d %H:%M:%S')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "BSECP模块数据"
        
        headers = [
            "模块ID", "模块代码", "模块名称", "产品ID", "产品代码", 
            "是否积分", "价格", "状态", "备注", "停用时间", "停用用户",
            "创建时间", "创建用户", "更新时间", "更新用户"
        ]
        
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        for row, module in enumerate(modules, 2):
            ws.cell(row=row, column=1, value=module['MD_ID'])
            ws.cell(row=row, column=2, value=module['MD_CODE'])
            ws.cell(row=row, column=3, value=module['MD_NAME'])
            ws.cell(row=row, column=4, value=module['MD_PRODUCTID'])
            ws.cell(row=row, column=5, value=module['MD_PRODUCTCODE'])
            ws.cell(row=row, column=6, value="是" if module['MD_ISPOINT'] == 1 else "否")
            ws.cell(row=row, column=7, value=float(module['MD_PRICE']) if module['MD_PRICE'] else 0)
            ws.cell(row=row, column=8, value="启用" if module['MD_STATE'] == 1 else "停用")
            ws.cell(row=row, column=9, value=module['MD_REMARK'])
            ws.cell(row=row, column=10, value=module['MD_FORBIT_DATE'])
            ws.cell(row=row, column=11, value=module['MD_FORBIT_USER'])
            ws.cell(row=row, column=12, value=module['MD_CREATE_DATE'])
            ws.cell(row=row, column=13, value=module['MD_CREATE_USER'])
            ws.cell(row=row, column=14, value=module['MD_MODIFY_DATE'])
            ws.cell(row=row, column=15, value=module['MD_MODIFY_USER'])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"BSECP模块导出-{timestamp}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        print(f"导出BSECP模块数据时出错: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': '导出失败'})

@bsecp_bp.route('/api/bsecp/authorization-stats')
@login_required
@permission_required('auth_query')
def get_bsecp_authorization_stats():
    try:
        range_type = request.args.get('range', 'week')
        
        connection = get_db_connection()
        with connection.cursor() as cursor:
            if range_type == 'week':
                days = 7
            elif range_type == 'halfmonth':
                days = 15
            elif range_type == 'month':
                days = 30
            else:
                days = 7
            
            cursor.execute("""
                SELECT 
                    DATE(CreateTime) as date,
                    SUM(CASE WHEN AutoAuthFlag = '自动化授权成功' THEN 1 ELSE 0 END) as success_count,
                    SUM(CASE WHEN AutoAuthFlag = '自动化授权失败' THEN 1 ELSE 0 END) as failed_count,
                    SUM(CASE WHEN AutoAuthFlag NOT IN ('自动化授权成功', '自动化授权失败') THEN 1 ELSE 0 END) as pending_count
                FROM OrderAutoAuthorizationQueue
                WHERE CreateTime >= DATE_SUB(CURDATE(), INTERVAL %s DAY)
                GROUP BY DATE(CreateTime)
                ORDER BY DATE(CreateTime) ASC
            """, (days,))
            
            results = cursor.fetchall()
        connection.close()
        
        dates = []
        success_data = []
        failed_data = []
        pending_data = []
        
        for row in results:
            if row['date']:
                dates.append(row['date'].strftime('%Y-%m-%d'))
            else:
                dates.append('未知')
            success_data.append(row['success_count'] or 0)
            failed_data.append(row['failed_count'] or 0)
            pending_data.append(row['pending_count'] or 0)
        
        return jsonify({
            'success': True,
            'dates': dates,
            'success_data': success_data,
            'failed_data': failed_data,
            'pending_data': pending_data
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e),
            'dates': [],
            'success_data': [],
            'failed_data': [],
            'pending_data': []
        })
