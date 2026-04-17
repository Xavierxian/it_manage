from flask import Blueprint, render_template, request, jsonify, send_file
from flask_login import login_required
from .database import get_db_connection
from .auth import permission_required
from .cache_manager import cache_manager
from . import config

physical_machines_bp = Blueprint('physical_machines', __name__)

cache = config.cache

@physical_machines_bp.route('/physical_machines')
@login_required
@permission_required('physical_machines')
def physical_machines():
    return render_template('physical_machines.html')

@physical_machines_bp.route('/api/xenserver/<int:xen_id>')
@login_required
@permission_required('physical_machines')
@cache.cached(timeout=300, key_prefix=lambda: cache_manager.get_full_key('PM_DETAIL', request.view_args.get('xen_id', 0)))
def get_xenserver_by_id(xen_id):
    try:
        connection = get_db_connection()
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM xenserver WHERE id=%s", (xen_id,))
            xenserver = cursor.fetchone()
        connection.close()
        
        if xenserver:
            return jsonify({'xenserver': xenserver})
        else:
            return jsonify({'error': '实体机不存在'}), 404
    except Exception as e:
        print(f"获取实体机详细信息时出错: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@physical_machines_bp.route('/api/xenserver')
@login_required
@permission_required('physical_machines')
@cache.cached(timeout=300, key_prefix=lambda: cache_manager.get_full_key('PM_LIST'))
def get_xenserver():
    try:
        connection = get_db_connection()
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM xenserver ORDER BY id")
            xens = cursor.fetchall()
        connection.close()
        
        return jsonify({'xenserver': xens})
    except Exception as e:
        print(f"获取实体机数据时出错: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'xenserver': []}), 500

@physical_machines_bp.route('/api/xenserver', methods=['POST'])
@login_required
def add_xenserver():
    try:
        data = request.get_json()
        
        connection = get_db_connection()
        with connection.cursor() as cursor:
            cursor.execute("""
                INSERT INTO xenserver (服务器IP, 型号, 购买途径, 购买日期, 端口, 登录密码, 内存, 磁盘, 硬盘类型, 内存已使用, 磁盘已使用, 内存剩余, 磁盘剩余, 剩余可开, 部门, 用途)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                data.get('服务器IP'),
                data.get('型号'),
                data.get('购买途径'),
                data.get('购买日期'),
                data.get('端口'),
                data.get('登录密码'),
                data.get('内存'),
                data.get('磁盘'),
                data.get('硬盘类型'),
                data.get('内存已使用'),
                data.get('磁盘已使用'),
                data.get('内存剩余'),
                data.get('磁盘剩余'),
                data.get('剩余可开'),
                data.get('部门'),
                data.get('用途')
            ))
            connection.commit()
        connection.close()
        
        cache.clear()
        
        return jsonify({'success': True, 'message': '实体机添加成功'})
    except Exception as e:
        print(f"添加实体机时出错: {e}")
        return jsonify({'success': False, 'message': '添加失败'})

@physical_machines_bp.route('/api/xenserver/<int:xen_id>', methods=['PUT'])
@login_required
def update_xenserver(xen_id):
    try:
        data = request.get_json()
        
        connection = get_db_connection()
        with connection.cursor() as cursor:
            cursor.execute("""
                UPDATE xenserver 
                SET 服务器IP=%s, 型号=%s, 购买途径=%s, 购买日期=%s, 端口=%s, 登录密码=%s, 内存=%s, 磁盘=%s, 硬盘类型=%s, 内存已使用=%s, 磁盘已使用=%s, 内存剩余=%s, 磁盘剩余=%s, 剩余可开=%s, 部门=%s, 用途=%s
                WHERE id=%s
            """, (
                data.get('服务器IP'),
                data.get('型号'),
                data.get('购买途径'),
                data.get('购买日期'),
                data.get('端口'),
                data.get('登录密码'),
                data.get('内存'),
                data.get('磁盘'),
                data.get('硬盘类型'),
                data.get('内存已使用'),
                data.get('磁盘已使用'),
                data.get('内存剩余'),
                data.get('磁盘剩余'),
                data.get('剩余可开'),
                data.get('部门'),
                data.get('用途'),
                xen_id
            ))
            connection.commit()
        connection.close()
        
        cache.clear()
        
        return jsonify({'success': True, 'message': '实体机更新成功'})
    except Exception as e:
        print(f"更新实体机时出错: {e}")
        return jsonify({'success': False, 'message': '更新失败'})

@physical_machines_bp.route('/api/xenserver/<int:xen_id>', methods=['DELETE'])
@login_required
def delete_xenserver(xen_id):
    try:
        connection = get_db_connection()
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM xenserver WHERE id=%s", (xen_id,))
            connection.commit()
        connection.close()
        
        from flask_caching import Cache
        cache.clear()
        
        return jsonify({'success': True, 'message': '实体机删除成功'})
    except Exception as e:
        print(f"删除实体机时出错: {e}")
        return jsonify({'success': False, 'message': '删除失败'})

@physical_machines_bp.route('/api/xenserver/export')
@login_required
@permission_required('physical_machines')
def export_xenserver():
    try:
        connection = get_db_connection()
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM xenserver ORDER BY id")
            xens = cursor.fetchall()
        connection.close()
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        import datetime
        
        wb = Workbook()
        ws = wb.active
        ws.title = "实体机数据"
        
        headers = ["ID", "服务器IP", "型号", "购买途径", "购买日期", "端口", "登录密码", "内存", "磁盘", "硬盘类型", "内存已使用", "磁盘已使用", "内存剩余", "磁盘剩余", "剩余可开", "部门", "用途"]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        for row_num, xen in enumerate(xens, 2):
            for col_num, key in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = xen.get(key, '')
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"实体机数据_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        print(f"导出实体机数据时出错: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
