from flask import Blueprint, render_template, request, jsonify, send_file
from flask_login import login_required
from .database import get_db_connection
from .auth import permission_required
from .cache_manager import cache_manager
from . import config
import datetime
from io import BytesIO

namespaces_bp = Blueprint('namespaces', __name__)

cache = config.cache

@namespaces_bp.route('/namespaces')
@login_required
@permission_required('namespaces')
def namespaces():
    return render_template('namespaces.html')

@namespaces_bp.route('/api/bseip')
@login_required
@permission_required('namespaces')
@cache.cached(timeout=300, key_prefix=lambda: cache_manager.get_full_key('NS_LIST'))
def get_bseip():
    try:
        connection = get_db_connection()
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM bseip ORDER BY 开通日期 DESC")
            bseips = cursor.fetchall()
        connection.close()
        
        for bseip in bseips:
            if bseip['开通日期']:
                bseip['开通日期'] = bseip['开通日期'].strftime('%Y-%m-%d')
            if bseip['到期日期']:
                bseip['到期日期'] = bseip['到期日期'].strftime('%Y-%m-%d')
        
        return jsonify({'bseip': bseips})
    except Exception as e:
        print(f"获取NameSpace数据时出错: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'bseip': []}), 500

@namespaces_bp.route('/api/bseip/<int:bseip_id>')
@login_required
@cache.cached(timeout=300, key_prefix=lambda: cache_manager.get_full_key('NS_DETAIL', request.view_args.get('bseip_id', 0)))
def get_bseip_by_id(bseip_id):
    try:
        connection = get_db_connection()
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM bseip WHERE id=%s", (bseip_id,))
            bseip = cursor.fetchone()
        connection.close()
        
        if bseip:
            if bseip['开通日期']:
                bseip['开通日期'] = bseip['开通日期'].strftime('%Y-%m-%d')
            if bseip['到期日期']:
                bseip['到期日期'] = bseip['到期日期'].strftime('%Y-%m-%d')
            return jsonify({'bseip': bseip})
        else:
            return jsonify({'error': 'NameSpace不存在'}), 404
    except Exception as e:
        print(f"获取NameSpace详细信息时出错: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@namespaces_bp.route('/api/bseip', methods=['POST'])
@login_required
def add_bseip():
    try:
        data = request.get_json()
        
        connection = get_db_connection()
        with connection.cursor() as cursor:
            cursor.execute("""
                INSERT INTO bseip (命名空间, 空间归属, 申请部门, 空间对接人, 服务工程师, 开通日期, 到期日期, 用途, 是否停用)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                data.get('命名空间'),
                data.get('空间归属'),
                data.get('申请部门'),
                data.get('空间对接人'),
                data.get('服务工程师'),
                data.get('开通日期'),
                data.get('到期日期'),
                data.get('用途'),
                data.get('是否停用')
            ))
            connection.commit()
        connection.close()
        
        cache.clear()
        
        return jsonify({'success': True, 'message': 'NameSpace添加成功'})
    except Exception as e:
        print(f"添加NameSpace时出错: {e}")
        return jsonify({'success': False, 'message': '添加失败'})

@namespaces_bp.route('/api/bseip/<int:bseip_id>', methods=['PUT'])
@login_required
def update_bseip(bseip_id):
    try:
        data = request.get_json()
        
        connection = get_db_connection()
        with connection.cursor() as cursor:
            cursor.execute("""
                UPDATE bseip 
                SET 命名空间=%s, 空间归属=%s, 申请部门=%s, 空间对接人=%s, 服务工程师=%s, 开通日期=%s, 到期日期=%s, 用途=%s, 是否停用=%s
                WHERE id=%s
            """, (
                data.get('命名空间'),
                data.get('空间归属'),
                data.get('申请部门'),
                data.get('空间对接人'),
                data.get('服务工程师'),
                data.get('开通日期'),
                data.get('到期日期'),
                data.get('用途'),
                data.get('是否停用'),
                bseip_id
            ))
            connection.commit()
        connection.close()
        
        cache.clear()
        
        return jsonify({'success': True, 'message': 'NameSpace更新成功'})
    except Exception as e:
        print(f"更新NameSpace时出错: {e}")
        return jsonify({'success': False, 'message': '更新失败'})

@namespaces_bp.route('/api/bseip/<int:bseip_id>', methods=['DELETE'])
@login_required
def delete_bseip(bseip_id):
    try:
        connection = get_db_connection()
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM bseip WHERE id=%s", (bseip_id,))
            connection.commit()
        connection.close()
        
        from flask_caching import Cache
        cache.clear()
        
        return jsonify({'success': True, 'message': 'NameSpace删除成功'})
    except Exception as e:
        print(f"删除NameSpace时出错: {e}")
        return jsonify({'success': False, 'message': '删除失败'})

@namespaces_bp.route('/api/bseip/export')
@login_required
@permission_required('namespaces')
def export_bseip():
    try:
        connection = get_db_connection()
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM bseip ORDER BY 开通日期 DESC")
            bseips = cursor.fetchall()
        connection.close()
        
        for bseip in bseips:
            if bseip['开通日期']:
                bseip['开通日期'] = bseip['开通日期'].strftime('%Y-%m-%d')
            if bseip['到期日期']:
                bseip['到期日期'] = bseip['到期日期'].strftime('%Y-%m-%d')
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        import datetime
        
        wb = Workbook()
        ws = wb.active
        ws.title = "NameSpace数据"
        
        headers = ["ID", "命名空间", "空间归属", "申请部门", "空间对接人", "服务工程师", "开通日期", "到期日期", "用途", "是否停用"]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        for row_num, bseip in enumerate(bseips, 2):
            for col_num, key in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = bseip.get(key, '')
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"NameSpace数据_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        print(f"导出NameSpace数据时出错: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
